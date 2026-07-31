"""Gestion des exports Excel LegalTech en français.

L'export est organisé en quatre feuilles lisibles :
- ``Synthèse`` : vue décisionnelle courte ;
- ``Analyse_détaillée`` : toutes les données d'analyse et de traçabilité ;
- ``Revue_humaine`` : champs à corriger/valider par les départements ;
- ``Légende`` : définition des scores, statuts et niveaux de confiance.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_EXCEL_DIR


EXCEL_EXPORT_VERSION = "excel_francais_v3"

BASE_COLUMNS = [
    "ID_Annonce", "Date_Email", "Nom_PDF", "Société", "Société_Source",
    "Catégorie", "Code_Événement_Principal", "Type_Événement_IA",
    "Événements_Secondaires", "Résumé_IA", "Montant_Capital_Original",
    "Capital_Normalisé_TND", "Nombre_Parts_Actions",
    "Valeur_Nominale_Originale", "Valeur_Nominale_TND",
    "Capital_Calculé_TND", "Écart_Capital_TND", "Statut_Contrôle_Capital",
    "Période_Souscription", "Version_Parseur_Montants",
    "Risques_Détectés_IA", "Preuves_Risques", "Évaluation_Risque_Dérivée",
    "Base_Évaluation_Risque", "Opportunités_Détectées_IA",
    "Preuves_Opportunités", "Opportunité_Potentielle_Dérivée",
    "Base_Opportunité_Dérivée", "Nature_Opportunité_Dérivée",
    "Score_Risque_IA", "Niveau_Risque_IA", "Détail_Score_Risque",
    "Score_Opportunité_IA", "Niveau_Opportunité_IA",
    "Détail_Score_Opportunité", "Portée_Score_Opportunité",
    "Action_Recommandée_IA", "Confiance_IA", "Confiance_Extraction",
    "Confiance_Interprétation", "Plafond_Confiance_Interprétation",
    "Raison_Plafond_Confiance", "Couverture_Preuves", "Scoring_Method",
    "Version_Scoring", "Version_Analyseur", "Version_Validation_Factuelle",
    "Statut_Validation_Factuelle", "Corrections_Automatiques",
    "Erreurs_Validation_Factuelle",
]

TRACEABILITY_COLUMNS = [
    "Index_Annonce", "Référence_Annonce", "Page_Début", "Page_Fin",
    "Nombre_Pages", "Caractères_Annonce", "Qualité_Extraction_Segment",
    "Méthode_Extraction", "Méthode_Segmentation", "Statut_Validation_PDF",
    "Statut_Validation_PDF_Global", "Statut_Validation_Annonce",
    "Avertissements_Validation", "JSON_Valide", "Erreur_LLM", "Modèle_LLM",
    "Extrait_Source",
]

CORRECTION_COLUMNS = [
    "Statut_Revue", "Société_Corrigée", "Type_Événement_Corrigé",
    "Résumé_Corrigé",
    "Risques_Corrigés", "Opportunités_Corrigées", "Score_Risque_Corrigé",
    "Score_Opportunité_Corrigé", "Action_Corrigée", "Commentaire_Humain",
    "Validateur", "Date_Validation",
]

EXPORT_COLUMNS = BASE_COLUMNS + TRACEABILITY_COLUMNS + CORRECTION_COLUMNS

FRENCH_HEADER_MAP = {
    "ID_Annonce": "Identifiant de l’annonce",
    "Date_Email": "Source d’import",
    "Nom_PDF": "Fichier PDF",
    "Société": "Nom de société proposé",
    "Société_Source": "Nom relevé dans le document",
    "Catégorie": "Catégorie métier",
    "Code_Événement_Principal": "Code de l’événement principal",
    "Type_Événement_IA": "Événement principal",
    "Événements_Secondaires": "Événements secondaires",
    "Résumé_IA": "Résumé factuel",
    "Montant_Capital_Original": "Capital publié (texte source)",
    "Capital_Normalisé_TND": "Capital normalisé (TND)",
    "Nombre_Parts_Actions": "Nombre de parts ou actions",
    "Valeur_Nominale_Originale": "Valeur nominale publiée",
    "Valeur_Nominale_TND": "Valeur nominale normalisée (TND)",
    "Capital_Calculé_TND": "Capital recalculé (TND)",
    "Écart_Capital_TND": "Écart de capital (TND)",
    "Statut_Contrôle_Capital": "Statut du contrôle du capital",
    "Période_Souscription": "Période de souscription",
    "Version_Parseur_Montants": "Version du parseur de montants",
    "Risques_Détectés_IA": "Risques explicitement mentionnés",
    "Preuves_Risques": "Preuves textuelles des risques",
    "Évaluation_Risque_Dérivée": "Évaluation de risque dérivée",
    "Base_Évaluation_Risque": "Base de l’évaluation du risque",
    "Opportunités_Détectées_IA": "Opportunités explicitement mentionnées",
    "Preuves_Opportunités": "Preuves textuelles des opportunités",
    "Opportunité_Potentielle_Dérivée": "Opportunité potentielle dérivée",
    "Base_Opportunité_Dérivée": "Base de l’opportunité dérivée",
    "Nature_Opportunité_Dérivée": "Nature de l’opportunité dérivée",
    "Score_Risque_IA": "Score de risque (0–100)",
    "Niveau_Risque_IA": "Niveau de risque",
    "Détail_Score_Risque": "Détail du score de risque",
    "Score_Opportunité_IA": "Score d’opportunité générique (0–100)",
    "Niveau_Opportunité_IA": "Niveau d’opportunité",
    "Détail_Score_Opportunité": "Détail du score d’opportunité",
    "Portée_Score_Opportunité": "Portée du score d’opportunité",
    "Action_Recommandée_IA": "Action recommandée",
    "Confiance_IA": "Confiance globale du pipeline",
    "Confiance_Extraction": "Confiance d’extraction",
    "Confiance_Interprétation": "Confiance d’interprétation",
    "Plafond_Confiance_Interprétation": "Plafond de confiance appliqué",
    "Raison_Plafond_Confiance": "Raison du plafond de confiance",
    "Couverture_Preuves": "Couverture des preuves",
    "Scoring_Method": "Méthode de calcul des scores",
    "Version_Scoring": "Version du moteur de scoring",
    "Version_Analyseur": "Version de l’analyseur",
    "Version_Validation_Factuelle": "Version de la validation factuelle",
    "Statut_Validation_Factuelle": "Statut de validation factuelle",
    "Corrections_Automatiques": "Corrections automatiques appliquées",
    "Erreurs_Validation_Factuelle": "Points à revoir manuellement",
    "Index_Annonce": "Numéro de l’annonce dans le PDF",
    "Référence_Annonce": "Référence officielle de l’annonce",
    "Page_Début": "Page de début",
    "Page_Fin": "Page de fin",
    "Nombre_Pages": "Nombre de pages",
    "Caractères_Annonce": "Nombre de caractères",
    "Qualité_Extraction_Segment": "Qualité d’extraction du segment",
    "Méthode_Extraction": "Méthode d’extraction",
    "Méthode_Segmentation": "Méthode de segmentation",
    "Statut_Validation_PDF": "Statut de validation du PDF",
    "Statut_Validation_PDF_Global": "Statut global du PDF",
    "Statut_Validation_Annonce": "Statut structurel de l’annonce",
    "Avertissements_Validation": "Avertissements structurels",
    "JSON_Valide": "Réponse JSON valide",
    "Erreur_LLM": "Erreur du modèle local",
    "Modèle_LLM": "Modèle local utilisé",
    "Extrait_Source": "Extrait du texte source",
    "Statut_Revue": "Statut de revue humaine",
    "Société_Corrigée": "Nom de société validé ou corrigé",
    "Type_Événement_Corrigé": "Événement corrigé",
    "Résumé_Corrigé": "Résumé corrigé",
    "Risques_Corrigés": "Risques corrigés",
    "Opportunités_Corrigées": "Opportunités corrigées",
    "Score_Risque_Corrigé": "Score de risque corrigé",
    "Score_Opportunité_Corrigé": "Score d’opportunité corrigé",
    "Action_Corrigée": "Action recommandée corrigée",
    "Commentaire_Humain": "Commentaire du validateur",
    "Validateur": "Nom du validateur",
    "Date_Validation": "Date de validation",
}

REVERSE_HEADER_MAP = {value: key for key, value in FRENCH_HEADER_MAP.items()}


def safe_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (list, tuple, set)):
        return "; ".join(safe_text(item) for item in value if safe_text(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value).strip()


def safe_number(value: Any, default: Any = 0) -> Any:
    try:
        return round(float(value), 3)
    except (TypeError, ValueError):
        return default


def get_first(analysis: dict, *keys: str, default: Any = "") -> Any:
    for key in keys:
        if key in analysis and analysis.get(key) not in (None, ""):
            return analysis.get(key)
    return default


def _evidence_text(items: Any) -> str:
    if not isinstance(items, list):
        return ""
    lines: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        description = safe_text(item.get("description"))
        proof = safe_text(item.get("preuve_textuelle"))
        if description and proof:
            lines.append(f"{description} — Preuve : « {proof} »")
    return "\n".join(lines)


def _breakdown_text(items: Any) -> str:
    if not isinstance(items, list):
        return ""
    lines: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        factor = safe_text(item.get("facteur"), "Facteur")
        points = item.get("points", 0)
        justification = safe_text(item.get("justification"))
        line = f"{points:+} — {factor}"
        if justification:
            line += f" — {justification}"
        lines.append(line)
    return "\n".join(lines)


def _validation_text(items: Any) -> str:
    if not isinstance(items, list):
        return ""
    lines: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        code = safe_text(item.get("code") or item.get("field"), "INFO")
        message = safe_text(
            item.get("message")
            or item.get("reason")
            or f"{item.get('before', '')} → {item.get('after', '')}"
        )
        lines.append(f"{code} : {message}")
    return "\n".join(lines)


def analysis_to_row(analysis: dict, pdf_name: str, index: int) -> dict:
    risk_score = get_first(analysis, "Score_Risque_IA", "score_risque", default=0)
    opportunity_score = get_first(analysis, "Score_Opportunité_IA", "score_opportunite", default=0)
    risk_level = get_first(analysis, "Niveau_Risque_IA", "niveau_risque", default="-")
    opportunity_level = get_first(analysis, "Niveau_Opportunité_IA", "niveau_opportunite", default="-")
    confidence = get_first(analysis, "Confiance_IA", "niveau_confiance", default=0)

    breakdown = analysis.get("score_breakdown", {})
    factual = analysis.get("validation_factuelle", {})
    financial = analysis.get("faits_financiers_valides", {})
    if not isinstance(breakdown, dict):
        breakdown = {}
    if not isinstance(factual, dict):
        factual = {}
    if not isinstance(financial, dict):
        financial = {}

    risks = get_first(analysis, "Risques_Détectés_IA", "risques_detectes", default=[])
    opportunities = get_first(analysis, "Opportunités_Détectées_IA", "opportunites_detectees", default=[])
    derived_risk = analysis.get("evaluation_risque_derivee", {})
    derived_opportunity = analysis.get("opportunite_potentielle_derivee", {})
    if not isinstance(derived_risk, dict):
        derived_risk = {}
    if not isinstance(derived_opportunity, dict):
        derived_opportunity = {}

    row = {
        "ID_Annonce": f"ANN-{index:04d}",
        "Date_Email": "Import PDF manuel",
        "Nom_PDF": pdf_name,
        "Société": safe_text(get_first(analysis, "Société", "societe", default="À vérifier")),
        "Société_Source": safe_text(analysis.get("societe_source")),
        "Catégorie": safe_text(get_first(analysis, "Catégorie", "categorie", default="À vérifier")),
        "Code_Événement_Principal": safe_text(analysis.get("type_evenement_code"), "AUTRE"),
        "Type_Événement_IA": safe_text(get_first(analysis, "Type_Événement_IA", "type_evenement", default="À classifier manuellement")),
        "Événements_Secondaires": safe_text(analysis.get("evenements_secondaires", [])),
        "Résumé_IA": safe_text(get_first(analysis, "Résumé_IA", "resume", default="")),
        "Montant_Capital_Original": safe_text(financial.get("capital_original")),
        "Capital_Normalisé_TND": safe_number(financial.get("capital_tnd"), ""),
        "Nombre_Parts_Actions": financial.get("share_count", "") if financial.get("share_count") is not None else "",
        "Valeur_Nominale_Originale": safe_text(financial.get("nominal_original")),
        "Valeur_Nominale_TND": safe_number(financial.get("nominal_tnd"), ""),
        "Capital_Calculé_TND": safe_number(financial.get("computed_capital_tnd"), ""),
        "Écart_Capital_TND": safe_number(financial.get("difference_tnd"), ""),
        "Statut_Contrôle_Capital": safe_text(financial.get("arithmetic_status"), "NON_APPLICABLE"),
        "Période_Souscription": safe_text(financial.get("subscription_period")),
        "Version_Parseur_Montants": safe_text(financial.get("parser_version"), "-"),
        "Risques_Détectés_IA": safe_text(risks),
        "Preuves_Risques": _evidence_text(analysis.get("risques_avec_preuves", [])),
        "Évaluation_Risque_Dérivée": safe_text(derived_risk.get("description")),
        "Base_Évaluation_Risque": safe_text(derived_risk.get("base")),
        "Opportunités_Détectées_IA": safe_text(opportunities),
        "Preuves_Opportunités": _evidence_text(analysis.get("opportunites_avec_preuves", [])),
        "Opportunité_Potentielle_Dérivée": safe_text(derived_opportunity.get("description")),
        "Base_Opportunité_Dérivée": safe_text(derived_opportunity.get("base")),
        "Nature_Opportunité_Dérivée": safe_text(derived_opportunity.get("nature")),
        "Score_Risque_IA": safe_number(risk_score, 0),
        "Niveau_Risque_IA": safe_text(risk_level, "-"),
        "Détail_Score_Risque": _breakdown_text(breakdown.get("risk_breakdown", [])),
        "Score_Opportunité_IA": safe_number(opportunity_score, 0),
        "Niveau_Opportunité_IA": safe_text(opportunity_level, "-"),
        "Détail_Score_Opportunité": _breakdown_text(breakdown.get("opportunity_breakdown", [])),
        "Portée_Score_Opportunité": safe_text(breakdown.get("opportunity_scope")),
        "Action_Recommandée_IA": safe_text(get_first(analysis, "Action_Recommandée_IA", "action_recommandee", default="Vérifier manuellement le document.")),
        "Confiance_IA": safe_number(confidence, 0),
        "Confiance_Extraction": safe_number(analysis.get("confiance_extraction"), 0),
        "Confiance_Interprétation": safe_number(analysis.get("confiance_interpretation"), 0),
        "Plafond_Confiance_Interprétation": safe_number(analysis.get("plafond_confiance_interpretation"), 0),
        "Raison_Plafond_Confiance": safe_text(analysis.get("raison_plafond_confiance"), "-"),
        "Couverture_Preuves": safe_text(analysis.get("statut_couverture_preuves") if analysis.get("couverture_preuves") is None else analysis.get("couverture_preuves")),
        "Scoring_Method": safe_text(get_first(analysis, "Scoring_Method", "scoring_method", default="-")),
        "Version_Scoring": safe_text(analysis.get("scoring_version"), "-"),
        "Version_Analyseur": safe_text(analysis.get("analyzer_version"), "-"),
        "Version_Validation_Factuelle": safe_text(factual.get("version"), "-"),
        "Statut_Validation_Factuelle": safe_text(factual.get("status"), "INCONNU"),
        "Corrections_Automatiques": _validation_text(factual.get("corrections", [])),
        "Erreurs_Validation_Factuelle": _validation_text(factual.get("issues", [])),
    }
    for column in TRACEABILITY_COLUMNS + CORRECTION_COLUMNS:
        row[column] = ""
    return row


def _ensure_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    result = dataframe.copy()
    for column in EXPORT_COLUMNS:
        if column not in result.columns:
            result[column] = ""
    return result[EXPORT_COLUMNS]


def _summary_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    result = pd.DataFrame({
        "Référence officielle": dataframe["Référence_Annonce"],
        "Fichier PDF": dataframe["Nom_PDF"],
        "Pages": dataframe.apply(
            lambda row: (
                str(row["Page_Début"])
                if row["Page_Début"] == row["Page_Fin"]
                else f"{row['Page_Début']}–{row['Page_Fin']}"
            ), axis=1,
        ),
        "Société": dataframe["Société"],
        "Catégorie métier": dataframe["Catégorie"],
        "Événement principal": dataframe["Type_Événement_IA"],
        "Résumé factuel": dataframe["Résumé_IA"],
        "Score de risque": dataframe["Score_Risque_IA"],
        "Niveau de risque": dataframe["Niveau_Risque_IA"],
        "Évaluation du risque": dataframe["Évaluation_Risque_Dérivée"],
        "Score d’opportunité": dataframe["Score_Opportunité_IA"],
        "Niveau d’opportunité": dataframe["Niveau_Opportunité_IA"],
        "Opportunité potentielle": dataframe["Opportunité_Potentielle_Dérivée"],
        "Action recommandée": dataframe["Action_Recommandée_IA"],
        "Validation factuelle": dataframe["Statut_Validation_Factuelle"],
        "Contrôle financier": dataframe["Statut_Contrôle_Capital"].replace({
            "NOT_APPLICABLE": "NON APPLICABLE",
            "UNVERIFIED": "À VÉRIFIER",
            "MISMATCH": "INCOHÉRENCE",
            "PASS": "CONFORME",
        }),
        "Confiance du pipeline": dataframe["Confiance_IA"],
        "Statut de revue": dataframe["Statut_Revue"],
    })
    return result


def _review_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "ID_Annonce", "Nom_PDF", "Référence_Annonce", "Société_Source",
        "Société", "Catégorie", "Type_Événement_IA", "Résumé_IA",
        "Risques_Détectés_IA",
        "Opportunités_Détectées_IA", "Évaluation_Risque_Dérivée",
        "Opportunité_Potentielle_Dérivée", "Score_Risque_IA",
        "Score_Opportunité_IA", "Action_Recommandée_IA", "Extrait_Source",
    ] + CORRECTION_COLUMNS
    review = dataframe[columns].copy()
    review["Statut_Revue"] = review["Statut_Revue"].replace("", "À valider")
    return review.rename(columns=FRENCH_HEADER_MAP)


def _legend_dataframe() -> pd.DataFrame:
    rows = [
        ("Score de risque", "0–29 Faible ; 30–59 Moyen ; 60–79 Élevé ; 80–100 Critique."),
        ("Score d’opportunité", "Potentiel LegalTech générique. Il ne constitue pas une recommandation d’investissement UGFS."),
        ("Confiance du pipeline", "Indicateur de qualité d’extraction et d’interprétation ; ce n’est pas une probabilité de vérité juridique."),
        ("CONFORME / PASS", "Le contrôle déterministe a réussi."),
        ("INCOHÉRENCE / MISMATCH", "Les données publiées ne sont pas arithmétiquement cohérentes ; revue humaine obligatoire."),
        ("À VÉRIFIER / UNVERIFIED", "Une donnée nécessaire au contrôle manque."),
        ("NON APPLICABLE", "Le contrôle du capital n’est pas pertinent pour ce type d’événement."),
        ("Risque explicite", "Risque directement appuyé par une preuve textuelle de l’annonce."),
        ("Risque dérivé", "Évaluation prudente issue de règles déterministes ; elle n’est pas citée comme telle dans le document."),
        ("Opportunité potentielle", "Au maximum une déduction prudente, soumise à l’éligibilité et à la stratégie UGFS."),
        ("Nom relevé dans le document", "Libellé extrait du PDF sans validation humaine ; il peut contenir une confusion OCR."),
        ("Nom de société proposé", "Version nettoyée automatiquement pour faciliter la lecture, tout en conservant le libellé source."),
        ("Nom de société validé ou corrigé", "Champ à compléter par le département lors de la revue humaine."),
        ("Revue humaine", "Compléter les cellules jaunes de la feuille Revue_humaine puis réimporter le fichier dans l’application."),
    ]
    return pd.DataFrame(rows, columns=["Terme", "Définition"])


NAVY = "13294B"
BLUE = "0B5D86"
CYAN = "DDF3FC"
LIGHT_BLUE = "EAF3F8"
LIGHT_GREY = "F4F8FB"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
WHITE = "FFFFFF"
BORDER = Side(style="thin", color="D6E2EE")


def _style_table_sheet(
    worksheet,
    title: str,
    header_row: int = 3,
    editable_headers: set[str] | None = None,
) -> None:
    editable_headers = editable_headers or set()
    max_column = worksheet.max_column
    last_letter = get_column_letter(max_column)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    title_cell = worksheet.cell(1, 1)
    title_cell.value = title
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(color=WHITE, bold=True, size=15)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    subtitle = worksheet.cell(2, 1)
    subtitle.value = (
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — "
        "Les scores servent à la priorisation et nécessitent une revue humaine."
    )
    subtitle.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    subtitle.font = Font(color=NAVY, italic=True, size=10)
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.auto_filter.ref = f"A{header_row}:{last_letter}{worksheet.max_row}"
    worksheet.sheet_view.showGridLines = False

    header_names: dict[int, str] = {}
    for cell in worksheet[header_row]:
        header_names[cell.column] = str(cell.value or "")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=BORDER)
    worksheet.row_dimensions[header_row].height = 42

    text_keywords = (
        "Résumé", "Risque", "Opportunité", "Action", "Preuve", "Base",
        "Détail", "Commentaire", "Extrait", "Correction", "Définition",
    )
    narrow_keywords = ("Score", "Page", "Confiance", "Statut", "Nombre", "Écart")

    for col_idx in range(1, max_column + 1):
        header = header_names.get(col_idx, "")
        letter = get_column_letter(col_idx)
        if any(keyword in header for keyword in text_keywords):
            width = 42
        elif any(keyword in header for keyword in narrow_keywords):
            width = 16
        elif header in {"Société", "Nom de société proposé", "Nom relevé dans le document", "Nom de société validé ou corrigé", "Catégorie métier", "Événement principal", "Fichier PDF"}:
            width = 30
        else:
            width = min(max(len(header) + 3, 14), 28)
        worksheet.column_dimensions[letter].width = width

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 58
        alternate = PatternFill("solid", fgColor=LIGHT_GREY if row_idx % 2 == 0 else WHITE)
        for col_idx in range(1, max_column + 1):
            cell = worksheet.cell(row_idx, col_idx)
            header = header_names.get(col_idx, "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=BORDER)
            cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW) if header in editable_headers else alternate

    # Formats des valeurs de confiance et de scores.
    for col_idx, header in header_names.items():
        letter = get_column_letter(col_idx)
        if "Confiance" in header:
            for cell in worksheet[letter][header_row:]:
                cell.number_format = "0.0%"
        if header.startswith("Score"):
            worksheet.conditional_formatting.add(
                f"{letter}{header_row + 1}:{letter}{worksheet.max_row}",
                CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=PatternFill("solid", fgColor=LIGHT_RED)),
            )


def _style_legend(worksheet) -> None:
    _style_table_sheet(worksheet, "Légende et règles de lecture", header_row=3)
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 100
    for row in worksheet.iter_rows(min_row=4):
        row[0].font = Font(bold=True, color=NAVY)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def export_results(results: list[dict]) -> Path:
    OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_EXCEL_DIR / f"analyse_legaltech_fr_{timestamp}.xlsx"

    detailed = _ensure_columns(pd.DataFrame(results))
    summary = _summary_dataframe(detailed)
    review = _review_dataframe(detailed)
    detailed_fr = detailed.rename(columns=FRENCH_HEADER_MAP)
    legend = _legend_dataframe()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Synthèse", startrow=2)
        detailed_fr.to_excel(writer, index=False, sheet_name="Analyse_détaillée", startrow=2)
        review.to_excel(writer, index=False, sheet_name="Revue_humaine", startrow=2)
        legend.to_excel(writer, index=False, sheet_name="Légende", startrow=2)

        _style_table_sheet(writer.sheets["Synthèse"], "Synthèse décisionnelle LegalTech")
        _style_table_sheet(writer.sheets["Analyse_détaillée"], "Analyse détaillée et traçabilité")
        editable = {FRENCH_HEADER_MAP[column] for column in CORRECTION_COLUMNS}
        _style_table_sheet(
            writer.sheets["Revue_humaine"],
            "Revue humaine — compléter uniquement les cellules jaunes",
            editable_headers=editable,
        )
        _style_legend(writer.sheets["Légende"])

        writer.sheets["Synthèse"].sheet_properties.tabColor = BLUE
        writer.sheets["Analyse_détaillée"].sheet_properties.tabColor = NAVY
        writer.sheets["Revue_humaine"].sheet_properties.tabColor = "E6B800"
        writer.sheets["Légende"].sheet_properties.tabColor = "71839A"

    return output_path


def read_corrected_excel(uploaded_file) -> pd.DataFrame:
    """Lire la feuille de revue et restaurer les noms internes attendus."""
    workbook = pd.ExcelFile(uploaded_file)
    if "Revue_humaine" in workbook.sheet_names:
        dataframe = pd.read_excel(workbook, sheet_name="Revue_humaine", header=2)
    elif "Analyse_détaillée" in workbook.sheet_names:
        dataframe = pd.read_excel(workbook, sheet_name="Analyse_détaillée", header=2)
    else:
        dataframe = pd.read_excel(uploaded_file)

    dataframe = dataframe.rename(columns=REVERSE_HEADER_MAP)
    return dataframe
